# import json
# import requests
# from openpyxl import load_workbook

# filePath= ".//TestData/Sphere.xlsx"
# wb = load_workbook(filePath)
# ws = wb.active
# User = ws['F2'].value
# PW = ws['G2'].value

# # API endpoint URL for token generation
# token_url = 'https://sandbox.v2.nonprod.spherewms.com/api/v4/jwt/token'
# postOrders = "https://sandbox.v2.nonprod.spherewms.com/api/v4/jwta/ceva_asn/postOrders"
# # Parameters required to generate the token

# payload = {
#     'username':User,
#     'password': PW
# }
# print(payload)
# # Make a POST request to the API endpoint to get the token
# response = requests.post(token_url, json=payload)
# if response.status_code == 200:
#     # Extract the token from the response JSON
#     token = response.json()['token']

#     # Now make a POST request to the ASN endpoint using the obtained token
#     with open(".//PythonAPI//data.json", "r") as file:
#         data = json.load(file)
#     # Add environment parameter or header
#     headers = {
#         'Authorization': f'Bearer {token}',
#         'X-Environment': 'targetwms'
#     }
#     # Make a POST request to the ASN endpoint with the token, environment information, and JSON data
#     asn_response = requests.post(postOrders, headers=headers, json=data)
#     # Check the response
#     print(asn_response.text)
#     # print(asn_response.json())
#     # print(asn_response.headers)
#     # print(asn_response.content)
# else:
#     print('Error:', response.text)

# import json
# from openpyxl import load_workbook
# from datetime import datetime

# filePath= ".//TestData/Sphere.xlsx"
# wb = load_workbook(filePath)
# ws = wb.active

# hawb = ws['B2'].value
# cevaAppointment = ws['C2'].value
# licensePlate = ws['D2'].value
# arriveByDate = ws['E2'].value

# # Format arriveByDate to exclude time part
# if isinstance(arriveByDate, datetime):
#     arriveByDate_str = arriveByDate.strftime('%Y-%m-%d')
# else:
#     arriveByDate_str = arriveByDate  # In case it's not a datetime object

# print(arriveByDate_str)

# with open(".//PythonAPI//data.json", "r") as file:
#     data = json.load(file)

# data["hawb"] = hawb
# data["cevaAppointment"] = cevaAppointment
# data["pallets"][0]["licensePlate"] = licensePlate
# data["pallets"][0]["itemOrders"][0]["storeOrderQtys"][0]["arriveByDate"] = arriveByDate_str

# # Open the JSON file in write mode
# with open(".//PythonAPI//data.json", "w") as file:
#     # Write the updated data back to the JSON file
#     json.dump(data, file, indent=4)


from openpyxl import load_workbook
# from openpyxl.styles import NumberFormat

# Load the workbook and select the active worksheet
filePath = "./TestData/Sphere.xlsx"
wb = load_workbook(filePath)
ws = wb.active

# Retrieve values from specific cells
hawb = ws['B2'].value
cevaAppointment = ws['C2'].value
arriveByDate = ws['E2'].value
license = ws['D2'].value
license1=str(license)
print(license1)
# Ensure license is a number, increment, and convert to an integer
licensePlate = int(license1) + 1

print(str(license))
print(str(licensePlate))

# Update the cell value and format it as a number
ws['D2'].value = licensePlate
# Save the workbook
wb.save(filePath)
