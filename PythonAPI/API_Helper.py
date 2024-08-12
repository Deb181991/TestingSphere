import json
import requests
from openpyxl import load_workbook
from datetime import datetime

filePath = "./TestData/Sphere.xlsx"
wb = load_workbook(filePath)
ws = wb.active
hawb = ws['B2'].value
cevaAppointment = ws['C2'].value
arriveByDate = ws['E2'].value
with open("./TestData/PIPO_licensePlate.json", "r") as file:
    finaldata = json.load(file)
license_plate_number = int(finaldata['licensePlate'])
new_license_plate_number = license_plate_number + 1
license_plate= str(0)+str(0) + str(new_license_plate_number) 
print(license_plate)
data = {'licensePlate': new_license_plate_number}
with open("./TestData/PIPO_licensePlate.json", "w") as file:
    json.dump(data, file)

if isinstance(arriveByDate, datetime):
    arriveByDate_str = arriveByDate.strftime('%Y-%m-%d')
else:
    arriveByDate_str = arriveByDate

with open(".//PythonAPI//data.json", "r") as file:
    data = json.load(file)

data["hawb"] = hawb
data["cevaAppointment"] = cevaAppointment
data["pallets"][0]["licensePlate"] = license_plate
data["pallets"][0]["itemOrders"][0]["storeOrderQtys"][0]["arriveByDate"] = arriveByDate_str

with open(".//PythonAPI//data.json", "w") as file:
    json.dump(data, file, indent=4)

User = ws['F2'].value
PW = ws['G2'].value
url = ws['H2'].value

token_url = url+'api/v4/jwt/token'
postOrders = url+"api/v4/jwta/ceva_asn/postOrders"
payload = {
    'username':User,
    'password': PW
}

response = requests.post(token_url, json=payload)
if response.status_code == 200:
    token = response.json()['token']

    with open(".//PythonAPI//data.json", "r") as file:
        data = json.load(file)
    headers = {
        'Authorization': f'Bearer {token}',
        'X-Environment': 'targetwms'
    }
    asn_response = requests.post(postOrders, headers=headers, json=data)
    print(asn_response.text)
else:
    print('Error:', response.text)


