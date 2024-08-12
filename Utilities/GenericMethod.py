import json
import random
import string
import time
import urllib.request
import os
from openpyxl import load_workbook
import pandas as pd
from openpyxl.reader.excel import load_workbook
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from Logs import log_file
from Utilities.customLogger import LogGen
from datetime import datetime, timedelta

myLogger = LogGen.loggeen()
log = log_file.get_logs()


# action = ActionChains(driver)
# actions = ActionChains(browser)

class genericmethod:

    # @ initializing Driver
    def __init__(self, driver):
        self.driver = driver
        defaultTime = self.driver.implicitly_wait(60)

    # @ send keys to text field
    # @ element
    def type(self, element, text):
        element.clear()
        element.send_keys(text)
        myLogger.info("****** Enter Data ******")

    # @ send keys to text field
    # @ element
    def click(self, element):
        element.click()
        myLogger.info("****** Clicking Element ******")

    # @ send keys to text field
    # @ element
    def wait(self, TimeUnit):
        time.sleep(TimeUnit)
        myLogger.info("****** Waiting ******")

    # @ send keys to text field
    # @ element
    def navigateTo(self, url):
        myLogger.info("Navigating to " + url + "")
        self.driver.get(url)

    # @ get Text from element
    # @ element
    def getText(self, element):
        myLogger.info("************ Retrieving text *************")
        data = element.text
        # driver.log(data)
        myLogger.info(data)
        return data

    # @ get Value from text field
    # @ element

    def ScrollDown(self):
        self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
        # self.driver.execute_script("window.scrollTo(0, 500)")
        myLogger.info("****** scrolled down ******")

    def getValue(self, element):
        myLogger.info("************ Retrieving value *************")
        data = element.get_property('value')
        # driver.log(data)
        myLogger.info(data)
        return data

    # @ Waiting for element to be visible
    # @ element
    def waitForElementToBeVisible(self, element, elementname):
        myLogger.info("Waiting for " + elementname + " to be visible")
        myLogger.info("************* Waiting For Element *************")
        element.is_displayed(self.driver.implicitly_wait(10))
        assert True

    # @ Waiting for element to be Present
    # @ element
    def waitForElementToBePresent(self, element, elementname):
        myLogger.info("Waiting for " + elementname + " to be present")
        element.is_displayed(self.driver.implicitly_wait(60))
        assert True

    # @ Waiting for element to be click
    # @ element
    def waitForElementToBeClickable(self, element, elementname):
        myLogger.info("Waiting for " + elementname + " to be clickable")
        element.is_displayed
        assert True

    # @ Waiting for page to be Load
    # @ element
    def waitForPageToLoad(self):
        myLogger.info("waiting for page to load")
        WebDriverWait(self.driver, self.SERVER_TIMEOUT).until(
            lambda wd: self.driver.execute_script("return document.readyState") == 'complete',
            "Page taking too long to load"
        )

    # @ Click using Javascript
    # @ element
    def clickUsingJS(self, element):
        myLogger.info("clicking using JS")
        self.waitForPageToLoad()
        # self.waitForElementToBePresent(element,elementName)
        # myLogger.info("Clicking at " + elementName + "")
        self.waitForPageToLoad()
        self.driver.execute_script('arguments[0].click();', element)

    # @ get data from file
    def ReadData(self, filePath, object):
        file = open(filePath, "r")
        finaldata = json.load(file)
        data = finaldata[object]
        print(finaldata)
        print(data)
        return data

    def WriteData(self, filePath, object, value):
        data = {}
        data[object] = value
        j = json.dumps(data)
        with open(filePath, "w") as f:
            f.write(j)
        return value

    # @ Perform mouse move action and click the target element
    # @ element
    def MouseHover(self, element):
        self.action.move_to_element(element).click().perform()
        myLogger.info("****** MouseHover Click an Element ******")

    def getExeclData(self, file_path, SheetName, Rowno):
        # df = pd.read_excel('D:\Bestinet\TestData\Testcase.xlsx', sheet_name=SheetName)
        df = pd.read_excel(file_path, sheet_name=SheetName)
        data = df.iloc[Rowno]
        return data

    def clearfield(self, element):
        element.clear()

    def pricechart_ratecalculation(self, path, Ratechartcode, weight):
        df = pd.read_excel(path, sheet_name='Sheet1')
        print(Ratechartcode)
        print(weight)
        Rowcount = len(df.index)

        for i in range(0, Rowcount):
            if df['Weight'].iloc[i] >= weight:
                break
        value1 = df[Ratechartcode].iloc[i]
        value = round(value1, 2)
        return value

    def Upload(self, element, Path):
        element.send_keys(Path)

    # Wait for the presence of an element on a web page.
    # driver: WebDriver instance.
    # by: The By strategy (e.g., By.ID, By.XPATH, etc.).
    # value: The locator value.
    # timeout: Maximum time to wait for the element (default is 10 seconds).
    # :return: The WebElement once it's present.

    # def wait_for_element(self, locator_strategy, locator_value, timeout=10):
    #     try:
    #         element = WebDriverWait(self.driver, timeout).until(
    #             EC.presence_of_element_located((locator_strategy, locator_value))
    #         )
    #         return element
    #     except Exception as e:
    #         print(f"Element not found: {e}")
    #         return None

    def wait_for_element(self, element, timeout=30):
        try:
            locatorWait = WebDriverWait(self.driver, timeout).until(
                EC.presence_of_element_located(element)
            )
            return locatorWait
        except Exception as e:
            print(f"Element not found: {e}")
            return None

    def Read_From_Excel(self, file_path, CellName):
        wb = load_workbook(file_path)
        ws = wb.active
        return ws[CellName].value

    def write_Excel_Data(self, FilePath, CellName, Value):
        wb = load_workbook(FilePath)
        ws = wb.active
        ws[CellName].value = Value
        return wb.save(FilePath)

    def generate_random_character(self, length):
        characters = string.ascii_letters + string.digits
        return ''.join(random.choice(characters) for _ in range(length))
    
    def switchToTab(self, windowCount):
        self.driver.switch_to.window(self.driver.window_handles[windowCount])

    def NewTab(self):
        self.driver.execute_script('window.open()')

    def CloseTab(self):
        self.driver.close()
        
    def Refresh(self):
        self.driver.refresh()

    def Clear(self, element):
        element.clear()        

    def WindowCount(self):
        multi_window1 = self.driver.window_handles
        length1 = len(multi_window1)
        v1 = length1 - 1
        myLogger.info(v1)
        self.driver.switch_to.window(self.driver.window_handles[v1])
   
    def generate_random_domain(self):
        username_length = random.randint(5, 10)
        username = ''.join(random.choices(string.ascii_lowercase + string.digits, k=username_length))
        email = username + '.com'
        return email 

    def generate_random_User(self):
        username_length = random.randint(5, 10)
        username = ''.join(random.choices(string.ascii_lowercase + string.digits, k=username_length))
        # email = username + '.com'
        return username 

    def monday(self):
        today = datetime.today()
        days_ahead = 0 - today.weekday() + 7
        next_monday_date = today + timedelta(days=days_ahead)
        return next_monday_date.day
        
    def api(self):
        with open(".//PythonAPI//data.json", "r") as file:
                data = json.load(file)
        # print(data)
        data["hawb"] = "CCCC13"
        data["cevaAppointment"] = "47895"
        data["pallets"][0]["licensePlate"] = "10000040000000501"
        data["pallets"][0]["itemOrders"][0]["storeOrderQtys"][0]["arriveByDate"] = "2024-06-03"
        # print(data)
        # Open the JSON file in write mode
        with open(".//PythonAPI//data.json", "w") as file:
        # Write the updated data back to the JSON file
            json.dump(data, file, indent=4)